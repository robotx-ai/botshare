#!/usr/bin/env python3
"""
build-robot-catalog.py — unify the messy vendor robot workbook into one clean table.

Reads the 34-sheet "Shopify Robot Sheet.xlsx", keeps only the main-robot sheets,
normalizes their wildly different layouts into a single schema, extracts every
embedded product image, and writes:

    data/robot-catalog/RobotCatalog.xlsx   (one sheet, 10 columns)
    data/robot-catalog/robot-images/       (extracted images, brand_model.<ext>)

Design spec: docs/superpowers/specs/2026-06-10-robot-catalog-unify-design.md

Run:  PYTHONPATH=~/.local/lib/python3.12/site-packages python3 scripts/build-robot-catalog.py
"""
import os, re, sys, zipfile, shutil
from pathlib import Path
import openpyxl

# ---------------------------------------------------------------- paths
REPO = Path(__file__).resolve().parents[1]
SRC = os.environ.get("ROBOT_XLSX", "/mnt/c/Users/zhour/Downloads/Shopify Robot Sheet.xlsx")
OUT_DIR = REPO / "data" / "robot-catalog"
IMG_DIR = OUT_DIR / "robot-images"
OUT_XLSX = OUT_DIR / "RobotCatalog.xlsx"

COLUMNS = ["BrandName", "SKU", "Picture", "Product name", "Model",
           "Description", "MSRP (USD)", "ServiceCategory", "CapabilityTag", "NeedsReview"]

# ---------------------------------------------------------------- per-sheet adapters
# col indices are 0-based. `id` = the column that identifies a product (a new row
# in it starts a new product; blank = description continuation of the prior product).
# `pic` = the column embedded images are anchored to (None = no images on sheet).
# base_cap / base_cat = default tags for the sheet, refined by keyword overrides.
ADAPTERS = [
    # ---- Unitree -------------------------------------------------------------
    {"sheet": "Unitree G1 Basic", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 1, "name": None, "desc": 2, "spec": None, "msrp": 4, "sku": None,
     "pic": 0, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
    {"sheet": "G1 Edu series", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 2, "name": 1, "desc": 3, "spec": None, "msrp": 4, "sku": None,
     "pic": 0, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
    {"sheet": "A2 Series", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 1, "name": None, "desc": 2, "spec": None, "msrp": 3, "sku": None,
     "pic": 0, "base_cap": "quadruped", "base_cat": "Warehouse"},
    {"sheet": "A2-W series", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 1, "name": None, "desc": 2, "spec": None, "msrp": 3, "sku": None,
     "pic": 0, "base_cap": "quadruped", "base_cat": "Warehouse"},
    {"sheet": "R1Basic", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 1, "name": None, "desc": 2, "spec": None, "msrp": 3, "sku": None,
     "pic": 0, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
    {"sheet": "R1 EDU", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 2, "name": 1, "desc": 3, "spec": None, "msrp": 4, "sku": None,
     "pic": 0, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
    {"sheet": "Go2-W", "brand": "Unitree", "hdr": 4,
     "id": 0, "model": 1, "name": 0, "desc": 3, "spec": None, "msrp": 4, "sku": None,
     "pic": 2, "base_cap": "quadruped", "base_cat": "Showcase & Performance"},
    {"sheet": "Go2", "brand": "Unitree", "hdr": 4,
     "id": 0, "model": 1, "name": 0, "desc": 3, "spec": None, "msrp": 4, "sku": None,
     "pic": 2, "base_cap": "quadruped", "base_cat": "Showcase & Performance"},
    {"sheet": "B2 series", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 2, "name": 1, "desc": 3, "spec": None, "msrp": 4, "sku": None,
     "pic": 0, "base_cap": "industrial", "base_cat": "Warehouse"},
    {"sheet": "B2-W series", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 2, "name": 1, "desc": 3, "spec": None, "msrp": 4, "sku": None,
     "pic": 0, "base_cap": "industrial", "base_cat": "Warehouse"},
    {"sheet": "G1-D", "brand": "Unitree", "hdr": 1,
     "id": 1, "model": 2, "name": 1, "desc": 3, "spec": None, "msrp": 4, "sku": None,
     "pic": 0, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
    # ---- Agibot --------------------------------------------------------------
    {"sheet": "Agibot智元", "brand": "Agibot", "hdr": 1,
     "id": 1, "model": 1, "name": 0, "desc": 3, "spec": None, "msrp": 4, "sku": 5,
     "pic": 2, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
    # ---- Keenon (delivery / cleaning) ---------------------------------------
    {"sheet": "Keenon擎朗", "brand": "Keenon", "hdr": 1,
     "id": 0, "model": 0, "name": None, "desc": 2, "spec": 3, "msrp": 4, "sku": 5,
     "pic": 1, "base_cap": "delivery", "base_cat": "Restaurant"},
    # ---- Pudu (delivery / reception / cleaning) -----------------------------
    {"sheet": "Pudu普度", "brand": "Pudu", "hdr": 1,
     "id": 0, "model": 0, "name": None, "desc": 2, "spec": None, "msrp": 3, "sku": 4,
     "pic": 1, "base_cap": "delivery", "base_cat": "Restaurant"},
    # ---- Booster (brand inferred from spec; flagged for review) -------------
    {"sheet": "K1", "brand": "Booster", "hdr": 2, "brand_uncertain": True,
     "id": 0, "model": 0, "name": None, "desc": 1, "spec": None, "msrp": 2, "sku": None,
     "pic": None, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
    {"sheet": "T1", "brand": "Booster", "hdr": 2, "brand_uncertain": True,
     "id": 0, "model": 0, "name": None, "desc": 1, "spec": None, "msrp": 2, "sku": None,
     "pic": None, "base_cap": "humanoid", "base_cat": "Showcase & Performance"},
]
# Empty in source, intentionally skipped: "Noetix松延动力", "Zeroth乐享科技"

# capability -> service category, used by keyword overrides
CAP_TO_CAT = {
    "delivery": "Restaurant", "reception": "Restaurant",
    "cleaning": "Warehouse", "industrial": "Warehouse",
    "humanoid": "Showcase & Performance", "quadruped": "Showcase & Performance",
}

# ---------------------------------------------------------------- helpers
def s(v):
    return "" if v is None else str(v).strip()

def clean_text(t):
    """Collapse newlines/whitespace — for model & product names (not descriptions)."""
    return re.sub(r"\s+", " ", str(t).replace("\n", " ")).strip()

# rows that are not robots: accessories, spares, warranty line-items
ACCESSORY_RE = re.compile(
    r"(?i)(battery|charging|charger|charge|charing|docking|remote\s*control|bracket|warranty|"
    r"spare|adapter|protective|\bpad\b|\bcable\b|\bstation\b|\bboard\b|extended\s+warranty)")

def parse_msrp(v):
    """Return number_or_None. Accepts 31900.0, '21600 USD', '8,990';
    rejects expressions/words like 'Complimentary' or 'Sales Price*25%'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v > 0 else None
    txt = re.sub(r"(?i)\b(usd|us\$|dollars?)\b", "", str(v))
    if re.search(r"[A-Za-z]", txt):  # leftover letters => not a clean price
        return None
    m = re.search(r"[\d][\d,\.]*", txt)
    if not m:
        return None
    try:
        n = float(m.group(0).replace(",", ""))
        return n if n > 0 else None
    except ValueError:
        return None

def fmt_msrp(n):
    return int(n) if n is not None and float(n).is_integer() else n

def slug(*parts):
    raw = "-".join(p for p in parts if p)
    raw = re.sub(r"[^A-Za-z0-9]+", "-", raw).strip("-").lower()
    return raw or "item"

def classify(model, desc, ad):
    """Return (capability, category, confident?). Keyword overrides the sheet default."""
    text = f"{model} {desc}".lower()
    if "cleaning" in text or "clean robot" in text:
        return "cleaning", "Warehouse", True
    if "reception" in text:
        return "reception", "Restaurant", True
    if "delivery" in text:
        return "delivery", "Restaurant", True
    if "humanoid" in text:
        return "humanoid", "Showcase & Performance", True
    if "quadruped" in text or "wheel-leg" in text:
        cap = ad["base_cap"] if ad["base_cap"] in ("quadruped", "industrial") else "quadruped"
        return cap, ad["base_cat"], True
    # no confirming keyword -> fall back to sheet default, mark for review
    return ad["base_cap"], ad["base_cat"], False

# ---------------------------------------------------------------- image extraction
def build_image_index(xlsx_path):
    """Map each main sheet -> list of (row0, col0, media_member) from drawing XML."""
    z = zipfile.ZipFile(xlsx_path)
    def rd(p):
        try:
            return z.read(p).decode("utf-8", "ignore")
        except KeyError:
            return ""
    wbx = rd("xl/workbook.xml")
    wbrels = rd("xl/_rels/workbook.xml.rels")
    order = re.findall(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wbx)
    ridmap = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', wbrels))
    name2sheet = {nm: "xl/" + ridmap[rid] for nm, rid in order if rid in ridmap}

    index = {}
    for name, sheetfile in name2sheet.items():
        sidx = re.search(r"sheet(\d+)\.xml", sheetfile).group(1)
        srels = rd(f"xl/worksheets/_rels/sheet{sidx}.xml.rels")
        d = re.search(r'Target="\.\./(drawings/drawing\d+\.xml)"', srels)
        if not d:
            continue
        didx = re.search(r"drawing(\d+)\.xml", d.group(1)).group(1)
        dxml = rd(f"xl/drawings/drawing{didx}.xml")
        drels = rd(f"xl/drawings/_rels/drawing{didx}.xml.rels")
        relmap = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="\.\./(media/[^"]+)"', drels))
        anchors = re.findall(
            r"<xdr:(?:oneCellAnchor|twoCellAnchor).*?</xdr:(?:oneCellAnchor|twoCellAnchor)>",
            dxml, re.S)
        items = []
        for a in anchors:
            col = re.search(r"<xdr:from>.*?<xdr:col>(\d+)</xdr:col>", a, re.S)
            row = re.search(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", a, re.S)
            emb = re.search(r'r:embed="(rId\d+)"', a)
            if col and row and emb and emb.group(1) in relmap:
                items.append((int(row.group(1)), int(col.group(1)), "xl/" + relmap[emb.group(1)]))
        index[name] = items
    return z, index

def pick_image(images, pic_col, start_row1, end_row1):
    """First image anchored in [start_row1, end_row1) on the picture column (0-based rows in XML)."""
    if pic_col is None:
        return None
    for row0, col0, media in images:
        if col0 == pic_col and start_row1 - 1 <= row0 < end_row1 - 1:
            return media
    return None

# ---------------------------------------------------------------- main
def main():
    if not Path(SRC).exists():
        sys.exit(f"Source workbook not found: {SRC}\nSet ROBOT_XLSX=/path/to/file.xlsx")
    if IMG_DIR.exists():
        shutil.rmtree(IMG_DIR)
    IMG_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(SRC, data_only=True)
    zf, img_index = build_image_index(SRC)

    rows_out = []
    log = []
    for ad in ADAPTERS:
        name = ad["sheet"]
        if name not in wb.sheetnames:
            log.append(f"  ! sheet missing, skipped: {name}")
            continue
        ws = wb[name]
        grid = list(ws.iter_rows(min_row=1, values_only=True))
        images = img_index.get(name, [])
        n = len(grid)

        def cell(r1, ci):
            if ci is None or r1 - 1 >= n:
                return ""
            row = grid[r1 - 1]
            return s(row[ci]) if ci < len(row) else ""

        # find product start rows: identifier column non-empty, after the header
        data_start = ad["hdr"] + 1
        starts = [r for r in range(data_start, n + 1) if cell(r, ad["id"])]
        starts.append(n + 1)  # sentinel for last span

        count = 0
        for i in range(len(starts) - 1):
            r0, r_next = starts[i], starts[i + 1]
            msrp = parse_msrp(grid[r0 - 1][ad["msrp"]] if ad["msrp"] < len(grid[r0 - 1]) else None)
            if msrp is None:
                continue  # section header / non-product row (e.g. "Service Robot", "Software Platform")

            model = clean_text(cell(r0, ad["model"]))
            name_val = clean_text(cell(r0, ad["name"]) if ad["name"] is not None else "")
            if not model:
                model = name_val
            if not model and not name_val:
                continue  # empty row, not a product
            # accessory check ignores parentheticals like "(with Battery, with Charger)"
            # which describe a robot bundle's contents, not an accessory product
            acc_str = re.sub(r"\([^)]*\)", "", f"{model} {name_val}")
            if ACCESSORY_RE.search(acc_str):
                continue  # accessory / spare / warranty line — out of scope (main robots only)
            product_name = name_val or f"{ad['brand']} {model}".strip()

            # description = start row + continuation rows (identifier blank), desc + spec cols
            desc_parts = []
            for rr in range(r0, r_next):
                for col in (ad["desc"], ad["spec"]):
                    val = cell(rr, col)
                    if val:
                        desc_parts.append(val)
            description = "\n".join(dict.fromkeys(desc_parts))  # dedupe, keep order

            sku = cell(r0, ad["sku"]) if ad["sku"] is not None else ""

            cap, cat, _ = classify(model, description, ad)
            # Per-sheet tags are reliable; only the inferred Booster brand is genuinely uncertain.
            needs_review = bool(ad.get("brand_uncertain"))

            # image
            media = pick_image(images, ad["pic"], r0, r_next)
            picture = ""
            if media:
                ext = Path(media).suffix.lower() or ".png"
                fname = f"{slug(ad['brand'], model)}{ext}"
                dest = IMG_DIR / fname
                # avoid collisions across products with same slug
                k = 2
                while dest.exists() and dest.name != fname:
                    fname = f"{slug(ad['brand'], model)}-{k}{ext}"; dest = IMG_DIR / fname; k += 1
                with open(dest, "wb") as fh:
                    fh.write(zf.read(media))
                picture = f"robot-images/{fname}"

            rows_out.append({
                "BrandName": ad["brand"], "SKU": sku, "Picture": picture,
                "Product name": product_name, "Model": model, "Description": description,
                "MSRP (USD)": fmt_msrp(msrp), "ServiceCategory": cat,
                "CapabilityTag": cap, "NeedsReview": "TRUE" if needs_review else "FALSE",
            })
            count += 1
        log.append(f"  - {name}: {count} products, {len(images)} images on sheet")

    # write output workbook
    out = openpyxl.Workbook()
    sh = out.active
    sh.title = "Robots"
    sh.append(COLUMNS)
    for r in rows_out:
        sh.append([r[c] for c in COLUMNS])
    # light formatting: header bold, freeze, widths, wrap description
    from openpyxl.styles import Font, Alignment
    for c in sh[1]:
        c.font = Font(bold=True)
    sh.freeze_panes = "A2"
    widths = {"A": 12, "B": 22, "C": 30, "D": 26, "E": 22, "F": 60,
              "G": 12, "H": 22, "I": 14, "J": 12}
    for col, w in widths.items():
        sh.column_dimensions[col].width = w
    for row in sh.iter_rows(min_row=2):
        row[5].alignment = Alignment(wrap_text=True, vertical="top")  # Description
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out.save(OUT_XLSX)

    # summary
    print("Per-sheet extraction:")
    print("\n".join(log))
    print(f"\nTotal products: {len(rows_out)}")
    print(f"With image:     {sum(1 for r in rows_out if r['Picture'])}")
    print(f"NeedsReview:    {sum(1 for r in rows_out if r['NeedsReview'] == 'TRUE')}")
    print(f"\nWrote: {OUT_XLSX}")
    print(f"Images: {IMG_DIR}  ({len(list(IMG_DIR.glob('*')))} files)")

if __name__ == "__main__":
    main()
